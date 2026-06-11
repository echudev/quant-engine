"""
CAFCI Data Fetcher — Camara Argentina de Fondos Comunes de Inversion.

API publica para fondos comunes argentinos. Sin API key, sin autenticacion.
La API REST anterior (api.pub.cafci.org.ar/tipo-renta, /fondo/...,
/estadisticas/...) fue discontinuada en 2026-04 (HTTP 403 "Route not
allowed"). Reemplazada por 4 fuentes alternativas:

Endpoints disponibles:

  CATALOGO (GET):
    - catalogo                Catalogo completo: 1152 fondos, 4615 clases con
                              IDs, honorarios, sociedad gerente, tipo renta,
                              region, horizonte, moneda. ~2.7 MB. Se cachea por dia.

  DIARIO (GET):
    - diario                  Snapshot diario XLSX: VCP, patrimonio, market
                              share, variaciones (dia/mes/YTD/12m) para todas
                              las clases activas. ~900 KB. Se parsea a JSON.

  FICHA INDIVIDUAL (GET):
    - ficha FONDO_ID CLASE_ID Ficha markdown (defuddle): rendimientos TNA por
                              periodo (7d/1m/90d/180d/YTD/12m).
    - cartera FONDO_ID CLASE_ID Composicion de cartera embebida en HTML como
                              data-pie-chart-items-value (top activos + %).

  CONSULTAS (sobre catalogo + diario en cache):
    - buscar QUERY            Buscar fondos por nombre parcial. Devuelve ID,
                              clases con fees.
    - top CATEGORIA           Top 10 fondos por patrimonio en una categoria
                              (ej: "Mercado de Dinero Peso Argentina").
    - fondo FONDO_ID CLASE_ID Ficha completa "todo en uno": metadata + fees
                              (catalogo) + patrimonio/variaciones (diario) +
                              rendimientos (ficha) + composicion (cartera).
    - resolve QUERY           Resolver fondoId/claseId desde nombre parcial.

  COMBINADOS:
    - all                     Snapshot catalogo + diario (sin fichas).

Uso:
    py fetch_cafci.py catalogo
    py fetch_cafci.py catalogo -o catalogo.json       # guarda completo (2.7MB)
    py fetch_cafci.py diario
    py fetch_cafci.py ficha 304 308                   # 1810 Ahorro
    py fetch_cafci.py cartera 304 308
    py fetch_cafci.py buscar "ahorro"
    py fetch_cafci.py buscar "ieb estrategico"
    py fetch_cafci.py top "Mercado de Dinero Peso Argentina"
    py fetch_cafci.py top "Renta Variable Peso Argentina" --limit 20
    py fetch_cafci.py fondo 304 308                   # all-in-one
    py fetch_cafci.py resolve "ahorro"
    py fetch_cafci.py all -o cafci_snapshot.json

Cache:
    Los archivos catalogo.json y diario.json se cachean por dia en
    $TMP/cafci-{catalog|daily}-YYYY-MM-DD.json. Asi las queries que
    cruzan catalogo+diario solo pegan a la API una vez por dia.
"""

from __future__ import annotations

import argparse
import html as html_mod
import json
import logging
import os
import re
import sys
import tempfile
import time
from datetime import date, datetime
from typing import Any

import requests

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

log = logging.getLogger("cafci")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)

# ── URLs y headers ─────────────────────────────────────────────────────────

URL_CATALOG = "https://estadisticas.cafci.org.ar/consulta-de-fondos.json"
URL_DAILY_XLSX = "https://api.pub.cafci.org.ar/pb_get"
URL_FICHA = "https://defuddle.md/https://estadisticas.cafci.org.ar/fondos/{fondo_id}?clase={clase_id}"
URL_HTML = "https://estadisticas.cafci.org.ar/fondos/{fondo_id}?clase={clase_id}"

HEADERS_BASIC = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
}

# /pb_get historicamente requeria estos headers para no devolver 403. Al
# 2026-06 funciona sin ellos, pero los seguimos enviando por seguridad.
HEADERS_BROWSER = {
    **HEADERS_BASIC,
    "Origin": "https://www.cafci.org.ar",
    "Referer": "https://www.cafci.org.ar/",
}

# Cache files. Se regeneran cada dia automaticamente.
CACHE_DIR = tempfile.gettempdir()


def _cache_path(name: str) -> str:
    """Ruta del cache para 'name' usando la fecha de hoy."""
    today = date.today().strftime("%Y-%m-%d")
    return os.path.join(CACHE_DIR, f"cafci-{name}-{today}.json")


# ── HTTP helpers ───────────────────────────────────────────────────────────


def _get(url: str, headers: dict | None = None, timeout: int = 60) -> requests.Response:
    """GET con error handling."""
    r = requests.get(url, headers=headers or HEADERS_BASIC, timeout=timeout)
    r.raise_for_status()
    return r


# ── Helpers para parsear XLSX ──────────────────────────────────────────────


def _to_iso_date(v):
    """Convierte date/datetime/string a ISO YYYY-MM-DD."""
    if v is None or v == "":
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _to_num(v):
    """Convierte a float si es posible."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _to_str(v):
    """Convierte a string strippeado o None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _parse_xlsx(xlsx_path: str) -> dict:
    """Parsea el XLSX de /pb_get a JSON normalizado.

    Estructura de filas:
      0-6: header con titulo y metadatos
      7-8: encabezados de columnas
      9: vacia
      10: nombre de categoria (solo col 0)
      11+: filas de fondos hasta proxima categoria
    """
    if openpyxl is None:
        raise ImportError("Necesita 'openpyxl' (pip install openpyxl)")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    fondos = []
    current_cat = None
    fecha_reporte = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 9:
            continue
        nonempty = [(j, v) for j, v in enumerate(row) if v is not None and v != ""]
        if not nonempty:
            continue
        # Fila de categoria: solo col 0 populada, sin fecha ni vcp.
        if len(nonempty) == 1 and nonempty[0][0] == 0:
            cat = _to_str(nonempty[0][1])
            if cat:
                current_cat = cat
            continue
        nombre = _to_str(row[0])
        if not nombre:
            continue
        fecha = _to_iso_date(row[4])
        if fecha and fecha_reporte is None:
            fecha_reporte = fecha
        fondos.append({
            "nombre": nombre,
            "categoria": current_cat,
            "moneda": _to_str(row[1]),
            "region": _to_str(row[2]),
            "horizonte": _to_str(row[3]),
            "fecha": fecha,
            "vcp_actual": _to_num(row[5]),
            "vcp_anterior": _to_num(row[6]),
            "variacion_dia_pct": _to_num(row[7]),
            "vcp_reexp_pesos": _to_num(row[8]),
            "variacion_mes_pct": _to_num(row[9]),
            "variacion_ytd_pct": _to_num(row[10]),
            "variacion_12m_pct": _to_num(row[11]),
            "cantidad_cuotapartes": _to_num(row[12]),
            "patrimonio": _to_num(row[14]),
            "market_share": _to_num(row[16]),
            "depositaria": _to_str(row[17]),
            "codigo_cnv": _to_str(row[18]),
        })

    seen = set()
    categorias = []
    for f in fondos:
        c = f["categoria"]
        if c and c not in seen:
            seen.add(c)
            categorias.append(c)

    return {
        "fecha_reporte": fecha_reporte,
        "categorias": categorias,
        "fondos": fondos,
    }


# ── Endpoints ──────────────────────────────────────────────────────────────


def fetch_catalogo(use_cache: bool = True) -> dict:
    """Catalogo completo de fondos + clases + IDs + honorarios + metadata.

    El catalogo tiene ~1152 fondos y ~4615 clases. ~2.7 MB.
    Se cachea por dia en $TMP/cafci-catalog-YYYY-MM-DD.json.

    Args:
        use_cache: si True (default), reusa el cache del dia si existe.

    Returns:
        dict con keys: generated_at, total_fondos, total_clases, filtros, fondos[].
        Ver REFERENCE.md para schema completo de fondos[].
    """
    path = _cache_path("catalog")
    if use_cache and os.path.exists(path) and os.path.getsize(path) > 0:
        log.info(f"Catalogo cargado de cache: {path}")
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    log.info(f"Descargando catalogo de {URL_CATALOG}...")
    r = _get(URL_CATALOG, timeout=120)
    data = r.json()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    log.info(f"Catalogo cacheado en {path} ({len(r.content)/1024:.1f} KB)")
    return data


def fetch_diario(use_cache: bool = True) -> dict:
    """Snapshot diario: VCP + patrimonio + market share + variaciones.

    Descarga el XLSX de /pb_get y lo parsea a JSON con `_parse_xlsx`.
    Cachea en $TMP/cafci-daily-YYYY-MM-DD.json.

    Args:
        use_cache: si True (default), reusa el cache del dia si existe.

    Returns:
        dict con keys: fecha_reporte (YYYY-MM-DD), categorias[], fondos[].
        Cada fondo: nombre, categoria, moneda, region, horizonte, fecha,
        vcp_actual, vcp_anterior, variacion_dia_pct, vcp_reexp_pesos,
        variacion_mes_pct, variacion_ytd_pct, variacion_12m_pct,
        cantidad_cuotapartes, patrimonio, market_share, depositaria, codigo_cnv.
    """
    path = _cache_path("daily")
    if use_cache and os.path.exists(path) and os.path.getsize(path) > 0:
        log.info(f"Diario cargado de cache: {path}")
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    log.info(f"Descargando XLSX diario de {URL_DAILY_XLSX}...")
    r = _get(URL_DAILY_XLSX, headers=HEADERS_BROWSER, timeout=120)
    xlsx_path = os.path.join(CACHE_DIR, f"cafci-daily-{date.today():%Y-%m-%d}.xlsx")
    with open(xlsx_path, "wb") as f:
        f.write(r.content)
    log.info(f"XLSX guardado en {xlsx_path} ({len(r.content)/1024:.1f} KB). Parseando...")

    parsed = _parse_xlsx(xlsx_path)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(parsed, f, ensure_ascii=False)
    log.info(f"Diario parseado y cacheado en {path}: {len(parsed['fondos'])} fondos, {len(parsed['categorias'])} categorias")
    return parsed


def fetch_ficha(fondo_id: int, clase_id: int) -> str:
    """Ficha individual del fondo en markdown (via defuddle.md).

    Contiene: rendimiento historico TNA (7d/1m/90d/180d/YTD/12m),
    valor cuotaparte actual, patrimonio bajo administracion, honorarios,
    datos del fondo (sociedad gerente/depositaria, tipo renta, region, etc.),
    inversion minima, plazo de liquidacion.

    Args:
        fondo_id: ID del fondo (ver catalogo).
        clase_id: ID de la clase (ver catalogo).

    Returns:
        string con el markdown de la ficha (~2 KB tipico).
    """
    url = URL_FICHA.format(fondo_id=fondo_id, clase_id=clase_id)
    log.info(f"Fetching ficha {fondo_id}/{clase_id}...")
    r = _get(url, timeout=30)
    return r.text


def fetch_cartera(fondo_id: int, clase_id: int) -> dict:
    """Composicion de cartera de un fondo+clase.

    Parsea el HTML de la pagina del fondo y extrae el atributo
    `data-pie-chart-items-value` del `<canvas>`, que contiene un JSON
    con los top activos (~14) y "Resto de Activos" agrupado.

    Args:
        fondo_id, clase_id: IDs del fondo y clase.

    Returns:
        dict con keys:
        - fecha_cartera: fecha de la cartera (formato DD/MM/YYYY).
        - composicion: lista de {nombre, porcentaje}.
    """
    url = URL_HTML.format(fondo_id=fondo_id, clase_id=clase_id)
    log.info(f"Fetching HTML cartera {fondo_id}/{clase_id}...")
    r = _get(url, headers=HEADERS_BROWSER, timeout=30)
    h = r.text

    m_date = re.search(r'class="valores">Valores al ([^<]+)<', h)
    m_pie = re.search(r'data-pie-chart-items-value="([^"]+)"', h)
    return {
        "fondo_id": fondo_id,
        "clase_id": clase_id,
        "fecha_cartera": m_date.group(1) if m_date else None,
        "composicion": json.loads(html_mod.unescape(m_pie.group(1))) if m_pie else [],
    }


# ── Consultas sobre cache ──────────────────────────────────────────────────


def buscar(query: str) -> list:
    """Buscar fondos por nombre parcial (case-insensitive) en el catalogo.

    Args:
        query: texto a buscar dentro del nombre del fondo.

    Returns:
        lista de fondos matched, cada uno con id, nombre, tipo_renta,
        sociedad_gerente, y resumen de clases (id, nombre, fees).
    """
    catalogo = fetch_catalogo()
    q = query.lower()
    matches = []
    for f in catalogo["fondos"]:
        if q in f["nombre"].lower():
            matches.append({
                "id": f["id"],
                "nombre": f["nombre"],
                "codigo_cnv": f.get("codigo_cnv"),
                "tipo_renta": f.get("tipo_renta", {}).get("nombre"),
                "moneda": f.get("moneda", {}).get("nombre"),
                "region": f.get("region", {}).get("nombre"),
                "sociedad_gerente": f.get("sociedad_gerente", {}).get("nombre"),
                "clases": [
                    {
                        "id": c["id"],
                        "nombre": c["nombre"],
                        "honorarios": c.get("honorarios", {}),
                        "inversion_minima": c.get("inversion_minima"),
                    }
                    for c in f.get("clases", [])
                ],
            })
    return matches


def resolve(query: str) -> list:
    """Resolver fondoId/claseId desde nombre parcial.

    Mas compacto que `buscar`. Devuelve solo IDs y nombres.

    Args:
        query: texto a buscar dentro del nombre del fondo.

    Returns:
        lista de {fondo_id, fondo_nombre, clase_id, clase_nombre}.
    """
    catalogo = fetch_catalogo()
    q = query.lower()
    out = []
    for f in catalogo["fondos"]:
        if q in f["nombre"].lower():
            for c in f.get("clases", []):
                out.append({
                    "fondo_id": f["id"],
                    "fondo_nombre": f["nombre"],
                    "clase_id": c["id"],
                    "clase_nombre": c["nombre"],
                })
    return out


def top(categoria: str, limit: int = 10) -> list:
    """Top N fondos por patrimonio en una categoria del diario XLSX.

    Las categorias combinan tipo_renta + moneda + region como string,
    ej: "Mercado de Dinero Peso Argentina", "Renta Variable Peso Argentina".
    Ver `diario.categorias[]` para la lista completa.

    Args:
        categoria: string exacto de la categoria.
        limit: cantidad a retornar (default 10).

    Returns:
        lista de fondos ordenados por patrimonio descendente.
    """
    diario = fetch_diario()
    matches = [f for f in diario["fondos"] if f["categoria"] == categoria]
    matches.sort(key=lambda f: f["patrimonio"] or 0, reverse=True)
    return matches[:limit]


def fondo_completo(fondo_id: int, clase_id: int) -> dict:
    """Ficha completa "todo en uno" combinando catalogo + diario + ficha + cartera.

    Devuelve un objeto unificado con:
      - meta: del catalogo (nombre, sociedad gerente/depositaria, tipo_renta, fees)
      - diario: del XLSX (patrimonio actual, variaciones)
      - ficha_md: markdown de defuddle (rendimientos por periodo)
      - cartera: composicion top activos

    Args:
        fondo_id, clase_id: IDs.

    Returns:
        dict con keys: meta, diario, ficha_md, cartera.
    """
    catalogo = fetch_catalogo()
    diario = fetch_diario()

    fondo = next((f for f in catalogo["fondos"] if f["id"] == fondo_id), None)
    if fondo is None:
        raise ValueError(f"fondo_id {fondo_id} no existe en catalogo")
    clase = next((c for c in fondo.get("clases", []) if c["id"] == clase_id), None)
    if clase is None:
        raise ValueError(f"clase_id {clase_id} no existe en fondo {fondo_id}")

    # Match diario por nombre exacto de la clase
    daily_match = next((d for d in diario["fondos"] if d["nombre"] == clase["nombre"]), None)

    # Las fichas markdown y cartera son request adicionales — los hacemos
    try:
        ficha_md = fetch_ficha(fondo_id, clase_id)
    except Exception as e:
        log.warning(f"ficha_md fallo: {e}")
        ficha_md = None
    try:
        cartera = fetch_cartera(fondo_id, clase_id)
    except Exception as e:
        log.warning(f"cartera fallo: {e}")
        cartera = None

    return {
        "meta": {
            "id": fondo["id"],
            "nombre": fondo["nombre"],
            "codigo_cnv": fondo.get("codigo_cnv"),
            "estado": fondo.get("estado"),
            "objetivo": fondo.get("objetivo"),
            "inicio": fondo.get("inicio"),
            "tipo_dinero": fondo.get("tipo_dinero"),
            "dias_liquidacion": fondo.get("dias_liquidacion"),
            "sociedad_gerente": fondo.get("sociedad_gerente", {}).get("nombre"),
            "sociedad_depositaria": fondo.get("sociedad_depositaria", {}).get("nombre"),
            "tipo_renta": fondo.get("tipo_renta", {}).get("nombre"),
            "region": fondo.get("region", {}).get("nombre"),
            "moneda": fondo.get("moneda", {}).get("nombre"),
            "horizonte": fondo.get("horizonte", {}).get("nombre"),
            "duration": fondo.get("duration", {}).get("nombre"),
            "benchmark": fondo.get("benchmark", {}).get("nombre"),
            "clase": {
                "id": clase["id"],
                "nombre": clase["nombre"],
                "inversion_minima": clase.get("inversion_minima"),
                "honorarios": clase.get("honorarios", {}),
                "ticker_bloomberg": (clase.get("ticker_bloomberg") or "").strip() or None,
                "ticker_isin": clase.get("ticker_isin"),
            },
        },
        "diario": daily_match,
        "ficha_md": ficha_md,
        "cartera": cartera,
    }


# ── Mode ALL ───────────────────────────────────────────────────────────────


def fetch_all() -> dict:
    """Snapshot del catalogo + diario. NO incluye fichas individuales."""
    log.info("Fetching ALL CAFCI snapshot...")
    result: dict[str, Any] = {"timestamp": datetime.now().isoformat()}
    try:
        result["catalogo"] = fetch_catalogo()
        time.sleep(0.3)
    except Exception as e:
        log.warning(f"catalogo: {e}")
        result["catalogo"] = {"error": str(e)}
    try:
        result["diario"] = fetch_diario()
    except Exception as e:
        log.warning(f"diario: {e}")
        result["diario"] = {"error": str(e)}
    return result


# ── CLI ────────────────────────────────────────────────────────────────────


MODES = [
    "catalogo", "diario",
    "ficha", "cartera",
    "buscar", "top", "fondo", "resolve",
    "all",
]


def main():
    parser = argparse.ArgumentParser(
        description="CAFCI Data Fetcher — Camara Argentina de Fondos Comunes de Inversion"
    )
    parser.add_argument("mode", choices=MODES, help="Modo de operacion")
    parser.add_argument(
        "args", nargs="*",
        help="Argumentos posicionales segun el modo (fondoId claseId, query, categoria)"
    )
    parser.add_argument(
        "--limit", type=int, default=10,
        help="Cantidad de resultados (modo top). Default: 10"
    )
    parser.add_argument(
        "--no-cache", action="store_true",
        help="Forzar refetch de catalogo/diario (ignorar cache local)"
    )
    parser.add_argument("-o", "--output", help="Guardar output a archivo")
    parser.add_argument("-q", "--quiet", action="store_true", help="Modo silencioso")

    args = parser.parse_args()
    if args.quiet:
        log.setLevel(logging.WARNING)

    use_cache = not args.no_cache

    try:
        if args.mode == "catalogo":
            result = fetch_catalogo(use_cache=use_cache)
        elif args.mode == "diario":
            result = fetch_diario(use_cache=use_cache)
        elif args.mode == "ficha":
            if len(args.args) < 2:
                log.error("ficha requiere FONDO_ID CLASE_ID")
                sys.exit(1)
            result = fetch_ficha(int(args.args[0]), int(args.args[1]))
        elif args.mode == "cartera":
            if len(args.args) < 2:
                log.error("cartera requiere FONDO_ID CLASE_ID")
                sys.exit(1)
            result = fetch_cartera(int(args.args[0]), int(args.args[1]))
        elif args.mode == "buscar":
            if not args.args:
                log.error("buscar requiere QUERY")
                sys.exit(1)
            result = buscar(" ".join(args.args))
        elif args.mode == "resolve":
            if not args.args:
                log.error("resolve requiere QUERY")
                sys.exit(1)
            result = resolve(" ".join(args.args))
        elif args.mode == "top":
            if not args.args:
                log.error("top requiere CATEGORIA (ej: 'Mercado de Dinero Peso Argentina')")
                sys.exit(1)
            result = top(" ".join(args.args), limit=args.limit)
        elif args.mode == "fondo":
            if len(args.args) < 2:
                log.error("fondo requiere FONDO_ID CLASE_ID")
                sys.exit(1)
            result = fondo_completo(int(args.args[0]), int(args.args[1]))
        elif args.mode == "all":
            result = fetch_all()
        else:
            parser.print_help()
            sys.exit(1)
    except requests.HTTPError as e:
        log.error(f"HTTP Error: {e}")
        if e.response is not None:
            log.error(f"Body: {e.response.text[:500]}")
        sys.exit(1)
    except Exception as e:
        log.error(f"Error: {e}")
        sys.exit(1)

    # ── Output ─────────────────────────────────────────────────────────
    if args.output:
        if isinstance(result, str):
            with open(args.output, "w", encoding="utf-8") as f:
                f.write(result)
        else:
            with open(args.output, "w", encoding="utf-8") as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
        log.info(f"Guardado: {args.output}")
    else:
        if isinstance(result, str):
            print(result)
        else:
            print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
